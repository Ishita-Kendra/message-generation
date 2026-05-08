import io, os, re, json, shutil, uuid, tempfile
from datetime import datetime
from flask import Flask, render_template, jsonify, request, send_file
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process as rfprocess

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

REFERENCE_DIR  = os.path.join(os.path.dirname(__file__), 'reference_files')
MANIFEST_PATH  = os.path.join(REFERENCE_DIR, '_manifest.json')
os.makedirs(REFERENCE_DIR, exist_ok=True)

# One-time download tokens  {token: (filepath, filename)}
_download_store = {}

OUTPUT_COLUMNS = [
    'Interested Person Contacted',
    'Interested Person Title',
    'Interested Person Location',
    'Interested Person Email',
    'Company',
    'New Contact Name',
    'New Contact Title',
    'New Contact Email',
    'New Contact Location',
    'Main Body Subject',
    'Main Body',
    'Follow-up Subject',
    'Follow-up Body',
    'Notes / Flags',
]

# ── Executive / title-level detection ────────────────────────────────────────

EXEC_KEYWORDS = {
    'president','ceo','cto','cfo','coo','cso','chief','founder','owner',
    'partner','chairman','board','executive','evp','svp','vice president',
    'director','head of','vp',
}
MGMT_KEYWORDS = {
    'manager','lead','coordinator','program','project','strategy',
    'operations','business development','account','commercial',
}

def contact_level(title):
    """Return 'exec', 'mgmt', or 'tech' based on title keywords."""
    if not title:
        return 'mgmt'
    t = title.lower()
    if any(k in t for k in EXEC_KEYWORDS):
        return 'exec'
    if any(k in t for k in MGMT_KEYWORDS):
        return 'mgmt'
    return 'tech'


# ── FEAAM message generation ──────────────────────────────────────────────────

def generate_subject(company, new_contact_title, contact_index):
    level = contact_level(new_contact_title)
    if contact_index == 0:
        return f"Continuing FEAAM's discussion at {company}"
    if level == 'exec':
        return f"Continuing FEAAM’s discussion at {company}"
    if level == 'tech':
        return f"Reduced rare-earth magnet mass for {company}’s applications"
    return f"FEAAM × {company} — parallel touchpoint"


def generate_main_body(new_first, new_title, company,
                       int_name, int_title, int_location, contact_index):
    level  = contact_level(new_title)
    level2 = contact_level(new_title)

    loc_bit = f", based in {int_location}," if int_location else ","
    title_bit = f", {int_title}" if int_title else ""
    role_bit = f"Given your role as {new_title}, the topic may be of interest." if new_title else \
               "The topic may be of interest at the strategic level."

    if level == 'exec' or contact_index == 0:
        body = (
            f"Hi {new_first},\n\n"
            f"We’ve recently been in contact with your colleague {int_name}{title_bit}"
            f"{loc_bit} regarding FEAAM’s patented stator flux barrier motor architecture, "
            f"which reduces rare-earth magnet mass while maintaining or improving torque density "
            f"and efficiency.\n\n"
            f"{role_bit}\n\n"
            f"Our architecture has been evaluated across applications ranging from material handling "
            f"to aerospace, supported by advanced electromagnetic simulation and optimization tools "
            f"developed in-house at FEAAM (a spin-off of Bundeswehr University Munich).\n\n"
            f"If useful, we’d be glad to share a complimentary Technical Analysis aligned with "
            f"{company}’s motor requirements.\n\n"
            f"Best regards,\nProf. Dr.-Ing. Dieter Gerling\nFounder, FEAAM GmbH"
        )
    elif level == 'mgmt':
        body = (
            f"Hi {new_first},\n\n"
            f"{int_name}{title_bit} (based in {int_location}), has recently engaged with us on "
            f"FEAAM’s patented stator flux barrier motor architecture, which reduces rare-earth "
            f"magnet mass while maintaining or improving torque density and efficiency under identical "
            f"electrical and geometrical boundary conditions.\n\n"
            f"In your role as {new_title}, the practical implications may be of direct relevance.\n\n"
            f"Our architecture has been evaluated across applications ranging from material handling "
            f"to aerospace, supported by advanced electromagnetic simulation and optimization tools "
            f"developed in-house at FEAAM (a spin-off of Bundeswehr University Munich).\n\n"
            f"If useful, we’d be glad to share a complimentary Technical Analysis aligned with "
            f"{company}’s motor requirements.\n\n"
            f"Best regards,\nProf. Dr.-Ing. Dieter Gerling\nFounder, FEAAM GmbH"
        )
    else:  # tech
        body = (
            f"Hi {new_first},\n\n"
            f"FEAAM’s patented stator flux barrier motor architecture reduces rare-earth magnet "
            f"mass while maintaining or improving torque density and efficiency under identical "
            f"electrical and geometrical boundary conditions.\n\n"
            f"Your colleague {int_name}{title_bit} has recently engaged with us on this topic. "
            f"Given your background in {new_title}, the architecture’s design and simulation "
            f"methodology may be of direct technical interest.\n\n"
            f"We’ve evaluated it across applications ranging from material handling to aerospace "
            f"and can share a complimentary Technical Analysis scoped to {company}’s requirements.\n\n"
            f"Best regards,\nProf. Dr.-Ing. Dieter Gerling\nFounder, FEAAM GmbH"
        )
    return body


def generate_followup(new_first, company, int_first, new_title, contact_index):
    level = contact_level(new_title)

    if level == 'exec' or contact_index == 0:
        body = (
            f"Hi {new_first},\n\n"
            f"Following up briefly on my earlier note. Given the parallel discussion with {int_first}, "
            f"the Technical Analysis can be reviewed alongside that thread without overlap.\n\n"
            f"Worth a brief look on your side?\n\n"
            f"Best regards,\nProf. Dr.-Ing. Dieter Gerling\nFounder, FEAAM GmbH"
        )
    elif level == 'mgmt':
        body = (
            f"Hi {new_first},\n\n"
            f"A short bump on my previous note. The Technical Analysis is intended as a low-effort "
            f"starting point for internal discussion and can be reviewed alongside {int_first}’s "
            f"thread without overlap.\n\n"
            f"Worth a brief look on your side?\n\n"
            f"Best regards,\nProf. Dr.-Ing. Dieter Gerling\nFounder, FEAAM GmbH"
        )
    else:  # tech
        body = (
            f"Hi {new_first},\n\n"
            f"Following up briefly. The architecture is patented and proven, with applications "
            f"evaluated across multiple drive contexts overlapping {company}’s portfolio.\n\n"
            f"Worth a brief look on your side?\n\n"
            f"Best regards,\nProf. Dr.-Ing. Dieter Gerling\nFounder, FEAAM GmbH"
        )
    return body


# ── Output reference loader ───────────────────────────────────────────────────

def load_output_reference():
    """
    Find the output-type reference file, read its column headers and all
    existing messages indexed by (company_key, contact_name_lower).
    Returns (columns_list, messages_dict).
    """
    manifest = load_manifest()
    for entry in sorted(manifest, key=lambda e: e.get('uploaded_at', ''), reverse=True):
        if entry.get('type') != 'output':
            continue
        path = os.path.join(REFERENCE_DIR, entry['name'])
        if not os.path.exists(path):
            continue
        try:
            # Detect header row (try rows 2, 1, 0)
            ref_df = None
            for hdr in [2, 1, 0]:
                try:
                    df = pd.read_excel(path, header=hdr, engine='openpyxl')
                    named = [c for c in df.columns if not str(c).startswith('Unnamed')]
                    if len(named) >= 5:
                        ref_df = df[named]
                        break
                except Exception:
                    continue
            if ref_df is None:
                continue

            cols = list(ref_df.columns)

            # Locate key columns (case-insensitive)
            def _fc(df, *kws):
                for c in df.columns:
                    cl = str(c).lower()
                    if all(k in cl for k in kws):
                        return c
                for c in df.columns:
                    cl = str(c).lower()
                    if any(k in cl for k in kws):
                        return c
                return None

            co_col      = _fc(ref_df, 'company')
            name_col    = _fc(ref_df, 'new contact name') or _fc(ref_df, 'new', 'name')
            subj_col    = _fc(ref_df, 'main body subject') or _fc(ref_df, 'main', 'subject')
            body_col    = _fc(ref_df, 'main body') or _fc(ref_df, 'main', 'body')
            fu_subj_col = _fc(ref_df, 'follow', 'subject')
            fu_body_col = _fc(ref_df, 'follow', 'body')

            messages = {}
            if co_col and name_col:
                for _, row in ref_df.iterrows():
                    co_key  = clean_company(safe_str(row.get(co_col, '')))
                    contact = safe_str(row.get(name_col, '')).lower().strip()
                    if not co_key or not contact:
                        continue
                    messages[(co_key, contact)] = {
                        'subject':    safe_str(row.get(subj_col,    '')) if subj_col    else '',
                        'body':       safe_str(row.get(body_col,    '')) if body_col    else '',
                        'fu_subject': safe_str(row.get(fu_subj_col, '')) if fu_subj_col else '',
                        'fu_body':    safe_str(row.get(fu_body_col, '')) if fu_body_col else '',
                    }

            return cols, messages

        except Exception as e:
            print(f'[output-ref] Could not load {entry["name"]}: {e}')

    return list(OUTPUT_COLUMNS), {}


# ── Manifest helpers ──────────────────────────────────────────────────────────

def load_manifest():
    if os.path.exists(MANIFEST_PATH):
        with open(MANIFEST_PATH, 'r') as f:
            return json.load(f)
    return []


def save_manifest(manifest):
    with open(MANIFEST_PATH, 'w') as f:
        json.dump(manifest, f, indent=2)


def save_to_reference(file_obj, original_name, file_type='reference'):
    manifest = load_manifest()
    safe_name = re.sub(r'[^\w.\-]', '_', original_name)
    dest_path = os.path.join(REFERENCE_DIR, safe_name)
    file_obj.seek(0)
    content = file_obj.read()
    file_obj.seek(0)
    with open(dest_path, 'wb') as f:
        f.write(content)
    # Preserve existing type if file already in manifest
    existing = next((e for e in manifest if e['name'] == safe_name), None)
    entry = {
        'name': safe_name,
        'original_name': original_name,
        'size': len(content),
        'uploaded_at': datetime.now().isoformat(timespec='seconds'),
        'type': existing['type'] if existing and 'type' in existing else file_type,
    }
    manifest = [e for e in manifest if e['name'] != safe_name]
    manifest.append(entry)
    save_manifest(manifest)
    return entry


# ── Reference file routes ─────────────────────────────────────────────────────

@app.route('/api/reference-files', methods=['GET'])
def list_reference_files():
    manifest = sorted(load_manifest(),
                      key=lambda e: e.get('uploaded_at', ''), reverse=True)
    return jsonify({'ok': True, 'files': manifest})


@app.route('/api/reference-files', methods=['POST'])
def upload_reference_file():
    files = request.files.getlist('files')
    if not files:
        return jsonify({'ok': False, 'error': 'No files provided'}), 400
    file_type = request.form.get('type', 'reference')
    if file_type not in ('reference', 'output'):
        file_type = 'reference'
    saved = [save_to_reference(f, f.filename, file_type) for f in files if f.filename]
    return jsonify({'ok': True, 'saved': saved})


@app.route('/api/reference-files/<filename>', methods=['DELETE'])
def delete_reference_file(filename):
    manifest = load_manifest()
    if not any(e['name'] == filename for e in manifest):
        return jsonify({'ok': False, 'error': 'File not found'}), 404
    path = os.path.join(REFERENCE_DIR, filename)
    if os.path.exists(path):
        os.remove(path)
    save_manifest([e for e in manifest if e['name'] != filename])
    return jsonify({'ok': True})


@app.route('/api/reference-files/<filename>', methods=['GET'])
def get_reference_file(filename):
    path = os.path.join(REFERENCE_DIR, filename)
    if not os.path.exists(path):
        return jsonify({'ok': False, 'error': 'File not found'}), 404
    return send_file(path, as_attachment=True)


# ── Column detection helpers ──────────────────────────────────────────────────

COMPANY_SUFFIXES = {
    'inc','llc','ltd','corp','co','gmbh','ag','sa','plc','limited',
    'incorporated','corporation','company','group','holding','holdings',
    'enterprises','solutions','services','international','global',
    'worldwide','consulting','the',
}


def clean_company(name):
    if not isinstance(name, str):
        return ''
    s = name.lower().strip()
    s = re.sub(r'[.,&\-/\\]+', ' ', s)
    return ' '.join(w for w in s.split() if w not in COMPANY_SUFFIXES).strip()


def find_col(df, *keywords):
    kws = [k.lower().replace(' ', '').replace('_', '') for k in keywords]
    for col in df.columns:
        cl = str(col).lower().replace(' ', '').replace('_', '').replace('-', '')
        if any(k in cl for k in kws):
            return col
    return None


def extract_name(row, df):
    for col in df.columns:
        cl = str(col).lower().strip()
        if cl in ('name', 'full name', 'fullname', 'contact name', 'contact'):
            v = str(row.get(col, '')).strip()
            if v and v.lower() != 'nan':
                return v
    fc = find_col(df, 'first name', 'firstname', 'first')
    lc = find_col(df, 'last name', 'lastname', 'last', 'surname')
    first = (str(row.get(fc, '')).strip() if fc else '').replace('nan', '').strip()
    last  = (str(row.get(lc, '')).strip() if lc else '').replace('nan', '').strip()
    if first or last:
        return f'{first} {last}'.strip()
    for col in df.columns:
        if 'name' in col.lower():
            v = str(row.get(col, '')).strip()
            if v and v.lower() != 'nan':
                return v
    return ''


def safe_str(v):
    s = str(v).strip()
    return '' if s.lower() == 'nan' else s


def read_excel(f):
    name = getattr(f, 'filename', '') or getattr(f, 'name', '')
    if str(name).lower().endswith('.csv'):
        try:
            return pd.read_csv(f)
        except Exception:
            f.seek(0)
            return pd.read_csv(f, encoding='latin-1')
    try:
        return pd.read_excel(f, engine='openpyxl')
    except Exception:
        f.seek(0)
        return pd.read_excel(f)


def read_excel_from_path(path):
    if path.lower().endswith('.csv'):
        try:
            return pd.read_csv(path)
        except Exception:
            return pd.read_csv(path, encoding='latin-1')
    try:
        return pd.read_excel(path, engine='openpyxl')
    except Exception:
        return pd.read_excel(path)


def open_file(src):
    if isinstance(src, str):
        return read_excel_from_path(os.path.join(REFERENCE_DIR, src))
    return read_excel(src)


def format_contacts(contacts):
    if not contacts:
        return ''
    if len(contacts) == 1:
        return contacts[0]
    if len(contacts) == 2:
        return f'{contacts[0]} and {contacts[1]}'
    return ', '.join(contacts[:-1]) + f', and {contacts[-1]}'


# ── Build detailed company map ────────────────────────────────────────────────

def build_company_map(df, co_col):
    """company_key → {display, contacts:[{name,title,location,email}]}"""
    title_col = find_col(df, 'title', 'position', 'role', 'designation', 'jobtitle')
    loc_col   = find_col(df, 'location', 'city', 'country', 'region', 'address')
    email_col = find_col(df, 'email', 'mail')

    co_map = {}
    for _, row in df.iterrows():
        co_raw = safe_str(row.get(co_col, ''))
        if not co_raw:
            continue
        co_key = clean_company(co_raw)
        if not co_key:
            continue
        name = extract_name(row, df)
        if not name:
            continue
        contact = {
            'name':     name,
            'title':    safe_str(row.get(title_col, '')) if title_col else '',
            'location': safe_str(row.get(loc_col,   '')) if loc_col   else '',
            'email':    safe_str(row.get(email_col,  '')) if email_col  else '',
        }
        if co_key not in co_map:
            co_map[co_key] = {'display': co_raw, 'contacts': []}
        if not any(c['name'] == name for c in co_map[co_key]['contacts']):
            co_map[co_key]['contacts'].append(contact)
    return co_map


def lookup_company(co_raw, co_map, fuzzy_thresh=80):
    co_key = clean_company(co_raw)
    if not co_key:
        return None, 'none'
    if co_key in co_map:
        return co_map[co_key], 'exact'
    if co_map:
        res = rfprocess.extractOne(co_key, list(co_map.keys()),
                                   scorer=fuzz.token_sort_ratio)
        if res and res[1] >= fuzzy_thresh:
            return co_map[res[0]], f'fuzzy ({res[1]}%)'
    return None, 'none'


# ── Core matching logic ───────────────────────────────────────────────────────

def build_output_rows(int_df, org_df, ref_messages=None):
    int_co_col = find_col(int_df, 'company','organization','org','employer','firm','account')
    org_co_col = find_col(org_df, 'company','organization','org','employer','firm','account')

    if not int_co_col:
        raise ValueError(f'No company column in Interested list. Found: {list(int_df.columns)}')
    if not org_co_col:
        raise ValueError(f'No company column in Organization list. Found: {list(org_df.columns)}')

    org_title_col = find_col(org_df, 'title','position','role','designation','jobtitle')
    org_loc_col   = find_col(org_df, 'location','city','country','region','address')
    org_email_col = find_col(org_df, 'email','mail')

    co_map = build_company_map(int_df, int_co_col)
    if ref_messages is None:
        ref_messages = {}

    # Track how many new contacts we've generated per company (for subject variation)
    company_contact_count = {}
    rows_out = []

    for _, row in org_df.iterrows():
        co_raw    = safe_str(row.get(org_co_col, ''))
        new_name  = extract_name(row, org_df)
        new_first = new_name.split()[0] if new_name else 'there'
        new_title = safe_str(row.get(org_title_col, '')) if org_title_col else ''
        new_email = safe_str(row.get(org_email_col, '')) if org_email_col else ''
        new_loc   = safe_str(row.get(org_loc_col,   '')) if org_loc_col   else ''

        entry, match_type = lookup_company(co_raw, co_map)

        if entry is None:
            rows_out.append({
                'int_name': '', 'int_title': '', 'int_location': '', 'int_email': '',
                'company': co_raw,
                'new_name': new_name, 'new_title': new_title,
                'new_email': new_email, 'new_location': new_loc,
                'subject': '', 'body': '', 'fu_subject': '', 'fu_body': '',
                'notes': '', 'match_type': 'none', 'has_match': False,
                'from_ref': False,
            })
            continue

        interested      = entry['contacts']
        company_display = co_raw or entry['display']
        primary         = interested[0]
        int_names_str   = format_contacts([c['name'] for c in interested])
        co_key          = clean_company(co_raw) or clean_company(entry['display'])
        idx             = company_contact_count.get(co_key, 0)
        company_contact_count[co_key] = idx + 1

        # ── Try to reuse messages from the output reference file ──────────────
        ref_key = (co_key, new_name.lower().strip())
        from_ref = False
        if ref_key in ref_messages and ref_messages[ref_key].get('body'):
            msgs     = ref_messages[ref_key]
            subject    = msgs['subject']
            main_body  = msgs['body']
            fu_subject = msgs['fu_subject']
            fu_body    = msgs['fu_body']
            from_ref   = True
        else:
            # Generate fresh messages following the reference style/templates
            subject    = generate_subject(company_display, new_title, idx)
            main_body  = generate_main_body(
                new_first, new_title, company_display,
                int_names_str, primary['title'], primary['location'], idx
            )
            fu_subject = f"Re: {subject}"
            fu_body    = generate_followup(
                new_first, company_display,
                primary['name'].split()[0], new_title, idx
            )

        rows_out.append({
            'int_name':     int_names_str,
            'int_title':    primary['title'],
            'int_location': primary['location'],
            'int_email':    primary['email'],
            'company':      company_display,
            'new_name':     new_name,
            'new_title':    new_title,
            'new_email':    new_email,
            'new_location': new_loc,
            'subject':      subject,
            'body':         main_body,
            'fu_subject':   fu_subject,
            'fu_body':      fu_body,
            'notes':        '✓ from reference' if from_ref else '',
            'match_type':   match_type,
            'has_match':    True,
            'from_ref':     from_ref,
        })

    return rows_out


def rows_to_df(rows_out, out_cols=None):
    """Convert output rows to a DataFrame using the reference column names."""
    if out_cols is None:
        out_cols = list(OUTPUT_COLUMNS)

    # Map our internal field names to the reference column names (flexible)
    # We match by keyword so it works even if the reference renames columns slightly.
    def pick_col(cols, *keywords):
        for kw in keywords:
            for c in cols:
                if kw.lower() in str(c).lower():
                    return c
        return None

    col_int_name  = pick_col(out_cols, 'interested person contacted', 'interested person', 'interested')
    col_int_title = pick_col(out_cols, 'interested person title')
    col_int_loc   = pick_col(out_cols, 'interested person location')
    col_int_email = pick_col(out_cols, 'interested person email')
    col_company   = pick_col(out_cols, 'company')
    col_new_name  = pick_col(out_cols, 'new contact name', 'new contact')
    col_new_title = pick_col(out_cols, 'new contact title')
    col_new_email = pick_col(out_cols, 'new contact email')
    col_new_loc   = pick_col(out_cols, 'new contact location')
    col_subj      = pick_col(out_cols, 'main body subject')
    col_body      = pick_col(out_cols, 'main body')
    col_fu_subj   = pick_col(out_cols, 'follow-up subject', 'followup subject', 'follow up subject')
    col_fu_body   = pick_col(out_cols, 'follow-up body', 'followup body', 'follow up body')
    col_notes     = pick_col(out_cols, 'notes', 'flags')

    data = []
    for r in rows_out:
        row = {}
        if col_int_name:  row[col_int_name]  = r['int_name']
        if col_int_title: row[col_int_title] = r['int_title']
        if col_int_loc:   row[col_int_loc]   = r['int_location']
        if col_int_email: row[col_int_email] = r['int_email']
        if col_company:   row[col_company]   = r['company']
        if col_new_name:  row[col_new_name]  = r['new_name']
        if col_new_title: row[col_new_title] = r['new_title']
        if col_new_email: row[col_new_email] = r['new_email']
        if col_new_loc:   row[col_new_loc]   = r['new_location']
        if col_subj:      row[col_subj]      = r['subject']
        if col_body:      row[col_body]      = r['body']
        if col_fu_subj:   row[col_fu_subj]   = r['fu_subject']
        if col_fu_body:   row[col_fu_body]   = r['fu_body']
        if col_notes:     row[col_notes]     = r['notes']
        data.append(row)

    return pd.DataFrame(data, columns=out_cols)


# ── Resolve input files ───────────────────────────────────────────────────────

def resolve_inputs(request, int_key='interested', org_key='organization'):
    int_name = request.form.get('interested_ref')
    org_name = request.form.get('organization_ref')
    int_f    = request.files.get(int_key)
    org_f    = request.files.get(org_key)

    if int_f and int_f.filename:
        save_to_reference(int_f, int_f.filename)
        int_f.seek(0)
    if org_f and org_f.filename:
        save_to_reference(org_f, org_f.filename)
        org_f.seek(0)

    int_src = int_name if int_name else int_f
    org_src = org_name if org_name else org_f

    if not int_src or not org_src:
        return None, None, 'Both an Interested List and an Organization List are required.'
    try:
        int_df = open_file(int_src)
        org_df = open_file(org_src)
    except Exception as e:
        return None, None, str(e)
    return int_df, org_df, None


# ── Message Generator routes ──────────────────────────────────────────────────

@app.route('/api/preview', methods=['POST'])
def preview():
    int_df, org_df, err = resolve_inputs(request)
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    out_cols, ref_messages = load_output_reference()
    try:
        rows_out = build_output_rows(int_df, org_df, ref_messages)
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

    matched   = sum(1 for r in rows_out if r['has_match'])
    from_ref  = sum(1 for r in rows_out if r.get('from_ref'))
    preview_rows = [{
        'int_name':      r['int_name'],
        'company':       r['company'],
        'new_name':      r['new_name'],
        'new_title':     r['new_title'],
        'subject':       r['subject'],
        'body_short':    r['body'][:120] + '…' if r['body'] else '',
        'fu_subject':    r['fu_subject'],
        'fu_body_short': r['fu_body'][:120] + '…' if r['fu_body'] else '',
        'match_type':    r['match_type'],
        'has_match':     r['has_match'],
        'from_ref':      r.get('from_ref', False),
    } for r in rows_out[:100]]

    return jsonify({'ok': True, 'total': len(rows_out),
                    'matched': matched, 'unmatched': len(rows_out) - matched,
                    'from_ref': from_ref,
                    'preview': preview_rows})


@app.route('/api/generate', methods=['POST'])
def generate():
    int_df, org_df, err = resolve_inputs(request)
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    out_cols, ref_messages = load_output_reference()
    try:
        rows_out = build_output_rows(int_df, org_df, ref_messages)
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

    out_df  = rows_to_df(rows_out, out_cols)
    matched = sum(1 for r in rows_out if r['has_match'])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # Title row
        out_df.to_excel(writer, index=False, startrow=2, sheet_name='FEAAM Outreach')
        ws = writer.sheets['FEAAM Outreach']

        # Row 1: title + summary
        title_text = (f"FEAAM Warm-Referral Outreach – Full Generation  |  "
                      f"Total warm leads: {matched}  |  Generated for new contacts at companies "
                      f"where an interested colleague exists.")
        ws.cell(row=1, column=1, value=title_text)
        ws.cell(row=1, column=1).font = Font(bold=True, size=11, color='E4E8F4')
        ws.cell(row=1, column=1).fill = PatternFill(start_color='111520',
                                                     end_color='111520', fill_type='solid')
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=len(out_cols))

        # Style header row (row 3)
        header_fill = PatternFill(start_color='1e2436', end_color='1e2436', fill_type='solid')
        header_font = Font(color='E4E8F4', bold=True, size=10)
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[3].height = 28

        # Highlight matched rows (light green), ref-reused rows (blue tint), unmatched (subtle)
        green_fill = PatternFill(start_color='0d2818', end_color='0d2818', fill_type='solid')
        blue_fill  = PatternFill(start_color='0d1a2e', end_color='0d1a2e', fill_type='solid')
        grey_fill  = PatternFill(start_color='161620', end_color='161620', fill_type='solid')
        for row_idx, r in enumerate(rows_out, start=4):
            if r.get('from_ref'):
                fill = blue_fill
            elif r['has_match']:
                fill = green_fill
            else:
                fill = grey_fill
            for col_idx in range(1, len(out_cols) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
            # Wrap text for body columns
            for col_name in out_cols:
                cl = col_name.lower()
                if 'body' in cl and 'subject' not in cl:
                    c_idx = out_cols.index(col_name) + 1
                    ws.cell(row=row_idx, column=c_idx).alignment = Alignment(wrap_text=True, vertical='top')

        # Column widths (defaults + overrides for known wide columns)
        default_widths = {
            'Interested Person Contacted': 28, 'Interested Person Title': 36,
            'Interested Person Location': 24,  'Interested Person Email': 28,
            'Company': 22, 'New Contact Name': 24, 'New Contact Title': 36,
            'New Contact Email': 28, 'New Contact Location': 20,
            'Main Body Subject': 44, 'Main Body': 55,
            'Follow-up Subject': 44, 'Follow-up Body': 55, 'Notes / Flags': 18,
        }
        for i, col in enumerate(out_cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = default_widths.get(col, 20)

        ws.freeze_panes = 'A4'

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    buf.seek(0)
    tmp.write(buf.read())
    tmp.close()
    token = uuid.uuid4().hex
    _download_store[token] = (tmp.name, 'FEAAM_Outreach_Messages.xlsx')
    return jsonify({'ok': True, 'token': token})


@app.route('/api/download/<token>')
def download_file(token):
    entry = _download_store.pop(token, None)
    if not entry:
        return 'Download link expired or invalid.', 404
    filepath, filename = entry
    if not os.path.exists(filepath):
        return 'File not found.', 404
    response = send_file(
        filepath,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    @response.call_on_close
    def _cleanup():
        try:
            os.unlink(filepath)
        except OSError:
            pass
    return response


# ── Deduplication routes ──────────────────────────────────────────────────────

def collect_other_keys(request):
    other_emails, other_names = set(), set()
    for f in request.files.getlist('others'):
        if not f.filename:
            continue
        save_to_reference(f, f.filename)
        f.seek(0)
        try:
            df = read_excel(f)
            _extract_keys(df, other_emails, other_names)
        except Exception:
            pass
    for name in request.form.getlist('others_ref'):
        try:
            df = read_excel_from_path(os.path.join(REFERENCE_DIR, name))
            _extract_keys(df, other_emails, other_names)
        except Exception:
            pass
    return other_emails, other_names


def _extract_keys(df, emails, names):
    for col in df.columns:
        cl = col.lower()
        if 'email' in cl or 'mail' in cl:
            vals = df[col].dropna().astype(str).str.lower().str.strip()
            emails.update(v for v in vals if v != 'nan')
        if 'name' in cl:
            vals = df[col].dropna().astype(str).str.lower().str.strip()
            names.update(v for v in vals if v != 'nan')


def _run_dedup(ref_df, other_emails, other_names):
    ref_email_col = find_col(ref_df, 'email', 'e-mail', 'mail')
    ref_name_col  = find_col(ref_df, 'full name', 'fullname', 'name')
    dup_rows, statuses, reasons = [], [], []
    for _, row in ref_df.iterrows():
        is_dup, reason = False, ''
        if ref_email_col:
            email = safe_str(row.get(ref_email_col, '')).lower()
            if email and email in other_emails:
                is_dup, reason = True, 'Email match'
        if not is_dup and ref_name_col:
            name = safe_str(row.get(ref_name_col, '')).lower()
            if name and name in other_names:
                is_dup, reason = True, 'Name match'
        statuses.append('Duplicate' if is_dup else 'Unique')
        reasons.append(reason)
        if is_dup:
            dup_rows.append({
                'name':   safe_str(row.get(ref_name_col, '')) if ref_name_col else '',
                'email':  safe_str(row.get(ref_email_col, '')) if ref_email_col else '',
                'reason': reason,
            })
    return statuses, reasons, dup_rows, ref_email_col, ref_name_col


def resolve_ref(request):
    ref_name = request.form.get('reference_ref')
    ref_f    = request.files.get('reference')
    if ref_f and ref_f.filename:
        save_to_reference(ref_f, ref_f.filename)
        ref_f.seek(0)
    src = ref_name if ref_name else ref_f
    if not src:
        return None, 'Reference file required'
    try:
        return open_file(src), None
    except Exception as e:
        return None, str(e)


@app.route('/api/dedup-preview', methods=['POST'])
def dedup_preview():
    ref_df, err = resolve_ref(request)
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    other_emails, other_names = collect_other_keys(request)
    statuses, _, dup_rows, _, _ = _run_dedup(ref_df, other_emails, other_names)
    return jsonify({'ok': True, 'total': len(ref_df),
                    'duplicates': len(dup_rows), 'unique': len(ref_df) - len(dup_rows),
                    'preview': dup_rows[:50]})


@app.route('/api/dedup-download', methods=['POST'])
def dedup_download():
    ref_df, err = resolve_ref(request)
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    other_emails, other_names = collect_other_keys(request)
    statuses, reasons, _, _, _ = _run_dedup(ref_df, other_emails, other_names)

    out_df = ref_df.copy()
    out_df['Status'] = statuses
    out_df['Match Reason'] = reasons

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        out_df.to_excel(writer, index=False, sheet_name='Deduplication')
        ws = writer.sheets['Deduplication']
        hf = PatternFill(start_color='111520', end_color='111520', fill_type='solid')
        for cell in ws[1]:
            cell.fill = hf
            cell.font = Font(color='E4E8F4', bold=True)
        red_fill = PatternFill(start_color='ffebee', end_color='ffebee', fill_type='solid')
        s_idx = list(out_df.columns).index('Status') + 1
        for r in range(2, len(out_df) + 2):
            if ws.cell(row=r, column=s_idx).value == 'Duplicate':
                for c in ws[r]:
                    c.fill = red_fill
                ws.cell(row=r, column=s_idx).font = Font(color='c62828', bold=True)
        for i in range(1, len(out_df.columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    buf.seek(0)
    tmp.write(buf.read())
    tmp.close()
    token = uuid.uuid4().hex
    _download_store[token] = (tmp.name, 'FEAAM_Deduplication.xlsx')
    return jsonify({'ok': True, 'token': token})


# ── Health / Index ────────────────────────────────────────────────────────────

@app.route('/api/health')
def health():
    return jsonify({'ok': True, 'service': 'FEAAM Contact Matcher'})


@app.route('/')
def index():
    return render_template('index.html')


if __name__ == '__main__':
    port = int(os.getenv('PORT', 5051))
    app.run(debug=True, port=port, host='0.0.0.0')
